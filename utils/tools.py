import matplotlib.pyplot as plt
from matplotlib_venn import venn2
import sys
# # 添加pyvenn路径
# sys.path.append(r'path\pyvenn-master')
import venn
from openpyxl import Workbook


# load lines from a file
def load_data(path):
    with open(path, 'r', encoding="UTF-8") as f:
        lines = f.read().split('\n')[0:-1]
    lines = [l.strip() for l in lines]
    return lines


def save_data(data, path):
    with open(path, "w", encoding='UTF-8') as f:
        for row in data:
            f.write(str(row) + '\n')


def draw_venn(DL_reserved_id, IR_reserved_id):
    if len(DL_reserved_id) == 0 or len(IR_reserved_id) == 0:
        print("Filter reserved NONE, maybe becasue the quality of dataset is too low or number of data is too few.")
    else:
        total = len(set(DL_reserved_id).union(set(IR_reserved_id)))
        # plt.style.use('seaborn-whitegrid')
        #  set_colors=['red', 'green']
        v = venn2([set(DL_reserved_id), set(IR_reserved_id)], set_labels=('DL Component', 'IR Component'), subset_label_formatter=lambda x: f"{(x/total):0.001%}")
        plt.savefig('./result_last/abalation_venn.jpg')
        plt.show()


def write_result_to_excel():
    # 创建一个Workbook对象
    wb = Workbook()
    # 获取当前活跃的sheet，默认是第一个sheet
    ws = wb.active

    ws['A1'] = 'class'
    ws['B1'] = 'name'
    ws['C1'].value = 'score'
    row1 = ['class1', 'zhangsan', 90]
    row2 = ['class2', 'lisi', 88]
    ws.append(row1)
    ws.append(row2)

    wb.save('data1.xlsx')
