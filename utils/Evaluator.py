import re
from decimal import Decimal
from random import sample
from nltk.translate.bleu_score import sentence_bleu
from nltk.translate.bleu_score import SmoothingFunction
from nltk.translate.meteor_score import meteor_score
from openpyxl import load_workbook, Workbook
from rouge import FilesRouge, Rouge
from utils.tools import load_data


def abalation_study_no_filter(test_title_path, test_pred_path):
    print("————————————————————————————————————————————————Abalation Study result————————————————————————————————————————————————")
    test_title = load_data(test_title_path)
    test_pred = load_data(test_pred_path)
    files_rouge = FilesRouge()
    raw_test_data_rouge_scores = files_rouge.get_scores(test_title_path, test_pred_path)
    average_bleu = 0
    average_meteor = 0
    average_rouge_1_f = 0
    average_rouge_2_f = 0
    average_rouge_l_f = 0
    for idx, title in enumerate(test_title):
        reference = []
        candiate = []
        for unit in title.split():
            for word in re.split("\!|\?|\,|\.|\'|\"|\{|\}|\`|\<|\>|\*|\@|\-|\#|\~|\$|\^|\+|\_", unit):
                if not word == '' and word not in reference:
                    reference.append(word)
        for unit in test_pred[idx].split():
            for word in re.split("\!|\?|\,|\.|\'|\"|\{|\}|\`|\<|\>|\*|\@|\-|\#|\~|\$|\^|\+|\_", unit):
                if not word == '' and word not in candiate:
                    candiate.append(word)
        smooth = SmoothingFunction()
        average_bleu += sentence_bleu([reference], candiate, weights=(0.25,0.25,0.25,0.25), smoothing_function=smooth.method1)
        average_meteor += meteor_score([reference], candiate)
    for score in raw_test_data_rouge_scores:
        rouge_1_f = score['rouge-1']['f']
        rouge_2_f = score['rouge-2']['f']
        rouge_l_f = score['rouge-l']['f']
        average_rouge_1_f += rouge_1_f
        average_rouge_2_f += rouge_2_f
        average_rouge_l_f += rouge_l_f
    average_bleu /= len(test_title)
    average_meteor /= len(test_title)
    average_rouge_1_f /= len(test_title)
    average_rouge_2_f /= len(test_title)
    average_rouge_l_f /= len(test_title)
    print("原始数据集 BLEU-4:{}, METEOR:{}, ROUGE-1-F1:{}, ROUGE-2-F1:{}, ROUGE-L-F1:{}".format(
        average_bleu, average_meteor, average_rouge_1_f, average_rouge_2_f, average_rouge_l_f))
    return average_bleu, average_meteor, average_rouge_1_f, average_rouge_2_f, average_rouge_l_f


def abalation_study_only_DL(test_title_path, test_pred_path, DL_reserved_id, save_path_DL_reserved_title, save_path_DL_reserved_pred):
    test_title = load_data(test_title_path)
    test_pred = load_data(test_pred_path)
    average_bleu = 0
    average_meteor = 0
    average_rouge_1_f = 0
    average_rouge_2_f = 0
    average_rouge_l_f = 0
    files_rouge = FilesRouge()
    DL_reserved_data_rouge_scores = files_rouge.get_scores(save_path_DL_reserved_title, save_path_DL_reserved_pred)
    for idx, id in enumerate(DL_reserved_id):
        reference = []
        candiate = []
        for unit in test_title[int(id)].split():
            for word in re.split("\!|\?|\,|\.|\'|\"|\{|\}|\`|\<|\>|\*|\@|\-|\#|\~|\$|\^|\+|\_", unit):
                if not word == '' and word not in reference:
                    reference.append(word)
        for unit in test_pred[int(id)].split():
            for word in re.split("\!|\?|\,|\.|\'|\"|\{|\}|\`|\<|\>|\*|\@|\-|\#|\~|\$|\^|\+|\_", unit):
                if not word == '' and word not in candiate:
                    candiate.append(word)
        smooth = SmoothingFunction()
        average_bleu += sentence_bleu([reference], candiate, weights=(0.25,0.25,0.25,0.25), smoothing_function=smooth.method1)
        average_meteor += meteor_score([reference], candiate)
    for score in DL_reserved_data_rouge_scores:
        rouge_1_f = score['rouge-1']['f']
        rouge_2_f = score['rouge-2']['f']
        rouge_l_f = score['rouge-l']['f']
        average_rouge_1_f += rouge_1_f
        average_rouge_2_f += rouge_2_f
        average_rouge_l_f += rouge_l_f
    DL_reserved_title = load_data(save_path_DL_reserved_title)
    average_bleu /= len(DL_reserved_title)
    average_meteor /= len(DL_reserved_title)
    average_rouge_1_f /= len(DL_reserved_title)
    average_rouge_2_f /= len(DL_reserved_title)
    average_rouge_l_f /= len(DL_reserved_title)
    print("DL模块过滤后 "
          "BLEU-4:{}, METEOR:{}, ROUGE-1-F1:{}, ROUGE-2-F1:{}, ROUGE-L-F1:{}".format(average_bleu, average_meteor, average_rouge_1_f, average_rouge_2_f, average_rouge_l_f))
    return average_bleu, average_meteor, average_rouge_1_f, average_rouge_2_f, average_rouge_l_f


def abalation_study_only_IR(test_title_path, test_pred_path, IR_reserved_id, save_path_IR_reserved_title, save_path_IR_reserved_pred):
    test_title = load_data(test_title_path)
    test_pred = load_data(test_pred_path)
    average_bleu = 0
    average_meteor = 0
    average_rouge_1_f = 0
    average_rouge_2_f = 0
    average_rouge_l_f = 0
    files_rouge = FilesRouge()
    IR_reserved_data_rouge_scores = files_rouge.get_scores(save_path_IR_reserved_title, save_path_IR_reserved_pred)
    for idx, id in enumerate(IR_reserved_id):
        reference = []
        candiate = []
        for unit in test_title[int(id)].split():
            for word in re.split("\!|\?|\,|\.|\'|\"|\{|\}|\`|\<|\>|\*|\@|\-|\#|\~|\$|\^|\+|\_", unit):
                if not word == '' and word not in reference:
                    reference.append(word)
        for unit in test_pred[int(id)].split():
            for word in re.split("\!|\?|\,|\.|\'|\"|\{|\}|\`|\<|\>|\*|\@|\-|\#|\~|\$|\^|\+|\_", unit):
                if not word == '' and word not in candiate:
                    candiate.append(word)
        smooth = SmoothingFunction()
        average_bleu += sentence_bleu([reference], candiate, weights=(0.25,0.25,0.25,0.25), smoothing_function=smooth.method1)
        average_meteor += meteor_score([reference], candiate)
    for score in IR_reserved_data_rouge_scores:
        rouge_1_f = score['rouge-1']['f']
        rouge_2_f = score['rouge-2']['f']
        rouge_l_f = score['rouge-l']['f']
        average_rouge_1_f += rouge_1_f
        average_rouge_2_f += rouge_2_f
        average_rouge_l_f += rouge_l_f
    IR_reserved_title = load_data(save_path_IR_reserved_title)
    average_bleu /= len(IR_reserved_title)
    average_meteor /= len(IR_reserved_title)
    average_rouge_1_f /= len(IR_reserved_title)
    average_rouge_2_f /= len(IR_reserved_title)
    average_rouge_l_f /= len(IR_reserved_title)
    print("IR模块过滤后 "
          "BLEU-4:{}, METEOR:{}, ROUGE-1-F1:{}, ROUGE-2-F1:{}, ROUGE-L-F1:{}".format(average_bleu, average_meteor, average_rouge_1_f, average_rouge_2_f, average_rouge_l_f))
    return average_bleu, average_meteor, average_rouge_1_f, average_rouge_2_f, average_rouge_l_f


def abalation_study_both(test_title_path, test_pred_path, both_reserved_id, save_path_both_reserved_title, save_path_both_reserved_pred):
    test_title = load_data(test_title_path)
    test_pred = load_data(test_pred_path)
    average_bleu = 0
    average_meteor = 0
    average_rouge_1_f = 0
    average_rouge_2_f = 0
    average_rouge_l_f = 0
    print("DL模块和IR模块协同过滤后保留的数据量：{}，占比约：{}%".format(len(both_reserved_id), float(Decimal((len(both_reserved_id) / len(test_title))).quantize(Decimal("0.0001"))) * 100))
    files_rouge = FilesRouge()
    reserved_data_rouge_scores = files_rouge.get_scores(save_path_both_reserved_title, save_path_both_reserved_pred)
    for idx, id in enumerate(both_reserved_id):
        reference = []
        candiate = []
        for unit in test_title[int(id)].split():
            for word in re.split("\!|\?|\,|\.|\'|\"|\{|\}|\`|\<|\>|\*|\@|\-|\#|\~|\$|\^|\+|\_", unit):
                if not word == '' and word not in reference:
                    reference.append(word)
        for unit in test_pred[int(id)].split():
            for word in re.split("\!|\?|\,|\.|\'|\"|\{|\}|\`|\<|\>|\*|\@|\-|\#|\~|\$|\^|\+|\_", unit):
                if not word == '' and word not in candiate:
                    candiate.append(word)
        smooth = SmoothingFunction()
        average_bleu += sentence_bleu([reference], candiate, weights=(0.25, 0.25, 0.25, 0.25), smoothing_function=smooth.method1)
        average_meteor += meteor_score([reference], candiate)
    for score in reserved_data_rouge_scores:
        rouge_1_f = score['rouge-1']['f']
        rouge_2_f = score['rouge-2']['f']
        rouge_l_f = score['rouge-l']['f']
        average_rouge_1_f += rouge_1_f
        average_rouge_2_f += rouge_2_f
        average_rouge_l_f += rouge_l_f
    both_reserved_title = load_data(save_path_both_reserved_title)
    average_bleu /= len(both_reserved_title)
    average_meteor /= len(both_reserved_title)
    average_rouge_1_f /= len(both_reserved_title)
    average_rouge_2_f /= len(both_reserved_title)
    average_rouge_l_f /= len(both_reserved_title)
    print("协同过滤后 "
          "BLEU-4:{}, METEOR:{}, ROUGE-1-F1:{}, ROUGE-2-F1:{}, ROUGE-L-F1:{}".format(average_bleu, average_meteor, average_rouge_1_f, average_rouge_2_f, average_rouge_l_f))
    both_num = len(both_reserved_id)
    both_rate = float(Decimal((len(both_reserved_id) / len(test_title))).quantize(Decimal("0.0001"))) * 100
    return average_bleu, average_meteor, average_rouge_1_f, average_rouge_2_f, average_rouge_l_f, both_num, both_rate


def compare_with_random(test_body_path, test_title_path, test_pred_path, both_reserved_rate):
    test_body = load_data(test_body_path)
    test_title = load_data(test_title_path)
    test_pred = load_data(test_pred_path)
    l = range(1, len(test_body))
    print("Random choose about {}% data from test set 10 times and the average performances:".format(both_reserved_rate*100))
    total_bleu = 0
    total_meteor = 0
    total_rouge_1_f = 0
    total_rouge_2_f = 0
    total_rouge_l_f = 0
    for i in range(0, 10):
        random_title = []
        random_pred = []
        random_list = sample(l, int(len(test_body) * both_reserved_rate))
        for id in random_list:
            random_title.append(test_title[id])
            random_pred.append(test_pred[id])
        average_bleu = 0
        average_meteor = 0
        average_rouge_1_f = 0
        average_rouge_2_f = 0
        average_rouge_l_f = 0
        for id in random_list:
            reference = []
            candiate = []
            for unit in test_title[id].split():
                for word in re.split("\!|\?|\,|\.|\'|\"|\{|\}|\`|\<|\>|\*|\@|\-|\#|\~|\$|\^|\+|\_", unit):
                    if not word == '' and word not in reference:
                        reference.append(word)
            for unit in test_pred[id].split():
                for word in re.split("\!|\?|\,|\.|\'|\"|\{|\}|\`|\<|\>|\*|\@|\-|\#|\~|\$|\^|\+|\_", unit):
                    if not word == '' and word not in candiate:
                        candiate.append(word)
            smooth = SmoothingFunction()
            average_bleu += sentence_bleu([reference], candiate, weights=(0.25, 0.25, 0.25, 0.25), smoothing_function=smooth.method1)
            average_meteor += meteor_score([reference], candiate)
        rouge = Rouge()
        for idx in range(len(random_title)):
            rouge_score = rouge.get_scores(random_pred[idx], random_title[idx])
            rouge_1_f = rouge_score[0]["rouge-1"]['f']
            rouge_2_f = rouge_score[0]["rouge-2"]['f']
            rouge_l_f = rouge_score[0]["rouge-l"]['f']
            average_rouge_1_f += rouge_1_f
            average_rouge_2_f += rouge_2_f
            average_rouge_l_f += rouge_l_f
        total_bleu += average_bleu
        total_meteor += average_meteor
        total_rouge_1_f += average_rouge_1_f
        total_rouge_2_f += average_rouge_2_f
        total_rouge_l_f += average_rouge_l_f
        average_bleu /= len(random_title)
        average_meteor /= len(random_title)
        average_rouge_1_f /= len(random_title)
        average_rouge_2_f /= len(random_title)
        average_rouge_l_f /= len(random_title)
        # print("\tBLEU-4:{}, Meteor:{}, Rouge-1:{}, Rouge-2:{}, Rouge-L:{}".format(average_bleu, average_meteor, average_rouge_1_f, average_rouge_2_f, average_rouge_l_f))
    total_bleu /= len(random_title) * 10
    total_meteor /= len(random_title) * 10
    total_rouge_1_f /= len(random_title) * 10
    total_rouge_2_f /= len(random_title) * 10
    total_rouge_l_f /= len(random_title) * 10
    print("\tBLEU-4:{}, Meteor:{}, Rouge-1:{}, Rouge-2:{}, Rouge-L:{}".format(average_bleu, average_meteor, average_rouge_1_f, average_rouge_2_f, average_rouge_l_f))
    return average_bleu, average_meteor, average_rouge_1_f, average_rouge_2_f, average_rouge_l_f


def evaluate_effectiveness_and_save_result_in_xlsx(model_name, test_body_path, test_title_path, test_pred_path,
                                                   DL_reserved_id, IR_reserved_id, save_path_DL_reserved_title,
                                                   save_path_DL_reserved_pred, save_path_IR_reserved_title,
                                                   save_path_IR_reserved_pred, save_path_both_reserved_title,
                                                   save_path_both_reserved_pred):
    test_title = load_data(test_title_path)
    DL_reserved_rate = float(
        Decimal((len(DL_reserved_id) / len(test_title))).quantize(Decimal("0.0001"), rounding="ROUND_HALF_UP")) * 100
    IR_reserved_rate = float(
        Decimal((len(IR_reserved_id) / len(test_title))).quantize(Decimal("0.0001"), rounding="ROUND_HALF_UP")) * 100
    print("DL module reserved: {}%, IR module reserved: {}%".format(DL_reserved_rate, IR_reserved_rate))

    both_reserved_id = list(set(DL_reserved_id) & set(IR_reserved_id))
    raw_average_bleu, raw_average_meteor, raw_average_rouge_1_f, raw_average_rouge_2_f, raw_average_rouge_l_f \
        = abalation_study_no_filter(test_title_path, test_pred_path)

    DL_average_bleu, DL_average_meteor, DL_average_rouge_1_f, DL_average_rouge_2_f, DL_average_rouge_l_f \
        = abalation_study_only_DL(test_title_path, test_pred_path, DL_reserved_id, save_path_DL_reserved_title,
                                  save_path_DL_reserved_pred)

    IR_average_bleu, IR_average_meteor, IR_average_rouge_1_f, IR_average_rouge_2_f, IR_average_rouge_l_f \
        = abalation_study_only_IR(test_title_path, test_pred_path, IR_reserved_id, save_path_IR_reserved_title,
                                  save_path_IR_reserved_pred)

    both_average_bleu, both_average_meteor, both_average_rouge_1_f, both_average_rouge_2_f, both_average_rouge_l_f, both_num, both_rate \
        = abalation_study_both(test_title_path, test_pred_path, both_reserved_id, save_path_both_reserved_title,
                               save_path_both_reserved_pred)

    random_average_bleu, random_average_meteor, random_average_rouge_1_f, random_average_rouge_2_f, random_average_rouge_l_f \
        = compare_with_random(test_body_path, test_title_path, test_pred_path, both_rate / 100)

    # 创建一个Workbook对象
    result = Workbook()
    # 获取当前活跃的sheet，默认是第一个sheet
    sheet1 = result.active

    # Table 3: Effectiveness of TitleGen-FL when applied to iTAPE based on the automatic evaluation
    sheet1.merge_cells('A1:G1')
    sheet1['A1'] = 'Table 3: Effectiveness of TitleGen-FL when applied to iTAPE based on the automatic evaluation'
    row_title = ['Approach', 'BLEU-4', 'METEOR', 'ROUGE-1', 'ROUGE-2', 'ROUGE-L', 'Filter Ratio']
    iTAPE = ['iTAPE', raw_average_bleu, raw_average_meteor, raw_average_rouge_1_f, raw_average_rouge_2_f,
             raw_average_rouge_l_f, 0]
    random = ['iTAPE+Random', random_average_bleu, random_average_meteor, random_average_rouge_1_f,
              random_average_rouge_2_f, random_average_rouge_l_f, (100 - both_rate) / 100]
    tool = ['iTAPE+tool', both_average_bleu, both_average_meteor, both_average_rouge_1_f, both_average_rouge_2_f,
            both_average_rouge_l_f, (100 - both_rate) / 100]
    sheet1.append(row_title)
    sheet1.append(iTAPE)
    sheet1.append(random)
    sheet1.append(tool)

    # Table 4: Effectiveness of TitleGen-FL when applied to iTAPE based on the automatic evaluation
    sheet1.merge_cells('A7:G7')
    sheet1['A7'] = 'Table 4: Effectiveness of TitleGen-FL when applied to iTAPE based on the automatic evaluation'
    row_title = ['Approach', 'BLEU-4', 'METEOR', 'ROUGE-1', 'ROUGE-2', 'ROUGE-L', 'Filter Ratio']
    DL = ['iTAPE+DL', DL_average_bleu, DL_average_meteor, DL_average_rouge_1_f, DL_average_rouge_2_f,
          DL_average_rouge_l_f, (100 - DL_reserved_rate) / 100]
    IR = ['iTAPE+IR', IR_average_bleu, IR_average_meteor, IR_average_rouge_1_f, IR_average_rouge_2_f,
          IR_average_rouge_l_f, (100 - IR_reserved_rate) / 100]
    tool = ['iTAPE+tool', both_average_bleu, both_average_meteor, both_average_rouge_1_f, both_average_rouge_2_f,
            both_average_rouge_l_f, (100 - both_rate) / 100]
    sheet1.append(row_title)
    sheet1.append(DL)
    sheet1.append(IR)
    sheet1.append(tool)

    human_study = load_workbook("Human Study for Filter.xlsx")
    human_study_sheet = human_study["Human Study"]
    DL_reserved_pred = load_data('./result_temp/DL_reserved_pred.txt')
    IR_reserved_pred = load_data('./result_temp/IR_reserved_pred.txt')

    # Body and pred
    body = []
    pred = []
    # Score list
    score = []
    score_DL = []
    score_IR = []
    score_both = []
    score_random = []
    # How many
    num_DL_res = 0
    num_IR_res = 0
    num_both_res = 0
    # How many high or low quality
    num_DL_high = 0
    num_DL_low = 0
    num_IR_high = 0
    num_IR_low = 0
    num_both_high = 0
    num_both_low = 0
    num_all_high = 0
    num_all_low = 0
    num_random_high = 0
    num_random_low = 0
    # Distribution of these scores
    dis_all = {1:0, 2:0, 3:0, 4:0, 5:0}
    dis_DL = {1:0, 2:0, 3:0, 4:0, 5:0}
    dis_IR = {1:0, 2:0, 3:0, 4:0, 5:0}
    dis_both = {1:0, 2:0, 3:0, 4:0, 5:0}
    dis_random = {1:0, 2:0, 3:0, 4:0, 5:0}
    # Sum of these scores
    sum_all = 0
    sum_DL = 0
    sum_IR = 0
    sum_both = 0
    sum_random = 0

    for i in human_study_sheet["A"]:
        if i.value == 'Body':
            continue
        else:
            body.append(i.value)
    for i in human_study_sheet["B"]:
        if i.value == 'Pred':
            continue
        else:
            pred.append(i.value)
    for i in human_study_sheet["C"]:
        if i.value == 'Score':
            continue
        else:
            score.append(int(Decimal(i.value).quantize(Decimal("1"))))

    # total count, high and low quality count, distribution for all, DL, IR and both
    for idx, unit_pred in enumerate(pred):
        if unit_pred in DL_reserved_pred:
            # get list
            score_DL.append(score[idx])
            # count total nums
            num_DL_res += 1
            # distribution
            if score[idx] in dis_DL.keys():
                dis_DL[score[idx]] += 1
            else:
                dis_DL[score[idx]] = 1
            # count high and low quality nums
            if score[idx] < 4 and score[idx] >= 0:
                num_DL_low += 1
            elif score[idx] >= 4 and score[idx] <= 5:
                num_DL_high += 1
            # calculate sum
            sum_DL += score[idx]
        if unit_pred in IR_reserved_pred:
            # get list
            score_IR.append(score[idx])
            # count total nums
            num_IR_res += 1
            # distribution
            if score[idx] in dis_IR.keys():
                dis_IR[score[idx]] += 1
            else:
                dis_IR[score[idx]] = 1
            # count high and low quality nums
            if score[idx] < 4 and score[idx] >= 0:
                num_IR_low += 1
            elif score[idx] >= 4 and score[idx] <= 5:
                num_IR_high += 1
            # calculate sum
            sum_IR += score[idx]
        if unit_pred in DL_reserved_pred and unit_pred in IR_reserved_pred:
            # get list
            score_both.append(score[idx])
            # count total nums
            num_both_res += 1
            # distribution
            if score[idx] in dis_both.keys():
                dis_both[score[idx]] += 1
            else:
                dis_both[score[idx]] = 1
            # count high and low quality nums
            if score[idx] < 4 and score[idx] >= 0:
                num_both_low += 1
            elif score[idx] >= 4 and score[idx] <= 5:
                num_both_high += 1
            # calculate sum
            sum_both += score[idx]
        # distribution
        if score[idx] in dis_all.keys():
            dis_all[score[idx]] += 1
        else:
            dis_all[score[idx]] = 1
        # count high and low quality nums
        if score[idx] < 4 and score[idx] >= 0:
            num_all_low += 1
        elif score[idx] >= 4 and score[idx] <= 5:
            num_all_high += 1
        # calculate sum
        sum_all += score[idx]

    # total count, high and low quality count, distribution for random
    l = range(1, len(score) + 1)
    choosed_number = sample(l, num_both_res)
    for num in choosed_number:
        # get list
        score_random.append(score[num - 1])
        # distribution
        if score[num - 1] in dis_random.keys():
            dis_random[score[num - 1]] += 1
        else:
            dis_random[score[num - 1]] = 1
        # count high and low quality nums
        if score[num - 1] < 4 and score[num - 1] >= 0:
            num_random_low += 1
        elif score[num - 1] >= 4 and score[num - 1] <= 5:
            num_random_high += 1
        # calculate sum
        sum_random += score[num - 1]

    # Table 5: Effectiveness of TitleGen-FL when applied to iTAPE based on the human study
    sheet1.merge_cells('A13:G13')
    sheet1['A13'] = 'Table 5: Effectiveness of TitleGen-FL when applied to iTAPE based on the human study'
    row_title = ['Approach', '1', '2', '3', '4', '5', 'Preserved', 'Mean Score']
    iTAPE = ['iTAPE', dis_all[1] / len(score), dis_all[2] / len(score), dis_all[3] / len(score), dis_all[4] / len(score),
             dis_all[5] / len(score), len(score), sum_all / len(score)]
    random = ['iTAPE+Random', dis_random[1] / num_both_res, dis_random[2] / num_both_res, dis_random[3] / num_both_res,
              dis_random[4] / num_both_res, dis_random[5] / num_both_res, num_both_res, sum_random / num_both_res]
    tool = ['iTAPE+tool', dis_both[1] / num_both_res, dis_both[2] / num_both_res, dis_both[3] / num_both_res,
            dis_both[4] / num_both_res, dis_both[5] / num_both_res, num_both_res, sum_both / num_both_res]
    sheet1.append(row_title)
    sheet1.append(iTAPE)
    sheet1.append(random)
    sheet1.append(tool)

    TP_random = num_random_high
    FP_random = num_random_low
    FN_random = num_all_high - num_random_high
    TN_random = num_all_low - num_random_low
    TP_both = num_both_high
    FP_both = num_both_low
    FN_both = num_all_high - num_both_high
    TN_both = num_all_low - num_both_low
    TP_DL = num_DL_high
    FP_DL = num_DL_low
    FN_DL = num_all_high - num_DL_high
    TN_DL = num_all_low - num_DL_low
    TP_IR = num_IR_high
    FP_IR = num_IR_low
    FN_IR = num_all_high - num_IR_high
    TN_IR = num_all_low - num_IR_low

    precisionHT_random = TP_random / (TP_random + FP_random)
    recallHT_random = TP_random / (TP_random + FN_random)
    HT_F1_random = 2 * precisionHT_random * recallHT_random / (precisionHT_random + recallHT_random)
    precisionHT_both = TP_both / (TP_both + FP_both)
    recallHT_both = TP_both / (TP_both + FN_both)
    HT_F1_both = 2 * precisionHT_both * recallHT_both / (precisionHT_both + recallHT_both)
    precisionHT_DL = TP_DL / (TP_DL + FP_DL)
    recallHT_DL = TP_DL / (TP_DL + FN_DL)
    HT_F1_DL = 2 * precisionHT_DL * recallHT_DL / (precisionHT_DL + recallHT_DL)
    precisionHT_IR = TP_IR / (TP_IR + FP_IR)
    recallHT_IR = TP_IR / (TP_IR + FN_IR)
    HT_F1_IR = 2 * precisionHT_IR * recallHT_IR / (precisionHT_IR + recallHT_IR)

    precisionLT_random = TN_random / (TN_random + FN_random)
    recallLT_random = TN_random / (TN_random + FP_random)
    LT_F1_random = 2 * precisionLT_random * recallLT_random / (precisionLT_random + recallLT_random)
    precisionLT_both = TN_both / (TN_both + FN_both)
    recallLT_both = TN_both / (TN_both + FP_both)
    LT_F1_both = 2 * precisionLT_both * recallLT_both / (precisionLT_both + recallLT_both)
    precisionLT_DL = TN_DL / (TN_DL + FN_DL)
    recallLT_DL = TN_DL / (TN_DL + FP_DL)
    LT_F1_DL = 2 * precisionLT_DL * recallLT_DL / (precisionLT_DL + recallLT_DL)
    precisionLT_IR = TN_IR / (TN_IR + FN_IR)
    recallLT_IR = TN_IR / (TN_IR + FP_IR)
    LT_F1_IR = 2 * precisionLT_IR * recallLT_IR / (precisionLT_IR + recallLT_IR)

    # Table 6: Effectiveness of TitleGen-FL in terms precision, recall and F1 when applied to iTAPE  based on the human study
    sheet1.merge_cells('A19:G19')
    sheet1[
        'A19'] = 'Table 6: Effectiveness of TitleGen-FL in terms precision, recall and F1 when applied to iTAPE based on the human study'
    row_title = ['Approach', 'PrecisionHT', 'RecallHT', 'F1HT', '#HT', 'PrecisionLT', 'RecallLT', 'F1LT', '#LT']
    random = ['iTAPE+Random', precisionHT_random, recallHT_random, HT_F1_random, num_both_res, precisionLT_random, recallLT_random, LT_F1_random, len(score) - num_both_res]
    both = ['iTAPE+tool', precisionHT_both, recallHT_both, HT_F1_both, num_both_res, precisionLT_both, recallLT_both, LT_F1_both, len(score) - num_both_res]
    sheet1.append(row_title)
    sheet1.append(random)
    sheet1.append(both)

    # Table 7: Effectiveness of TitleGen-FL when applied to iTAPE based on the human stud
    sheet1.merge_cells('A24:G24')
    sheet1['A24'] = 'Table 7: Effectiveness of TitleGen-FL when applied to iTAPE based on the human study'
    row_title = ['Approach', '1', '2', '3', '4', '5', 'Preserved', 'Mean Score']
    DL = ['iTAPE+DL', dis_DL[1] / num_DL_res, dis_DL[2] / num_DL_res, dis_DL[3] / num_DL_res, dis_DL[4] / num_DL_res,
          dis_DL[5] / num_DL_res, num_DL_res, sum_DL / num_DL_res]
    IR = ['iTAPE+IR', dis_IR[1] / num_IR_res, dis_IR[2] / num_IR_res, dis_IR[3] / num_IR_res, dis_IR[4] / num_IR_res,
          dis_IR[5] / num_IR_res, num_IR_res, sum_IR / num_IR_res]
    tool = ['iTAPE+tool', dis_both[1] / num_both_res, dis_both[2] / num_both_res, dis_both[3] / num_both_res,
            dis_both[4] / num_both_res, dis_both[5] / num_both_res, num_both_res, sum_both / num_both_res]
    sheet1.append(row_title)
    sheet1.append(DL)
    sheet1.append(IR)
    sheet1.append(tool)

    # Table 8: Effectiveness of TitleGen-FL in terms precision and recall when applied to iTAPE based on the human study
    sheet1.merge_cells('A30:G30')
    sheet1['A30'] = 'Table 8: Effectiveness of TitleGen-FL in terms precision and recall when applied to iTAPE based on the human study'
    row_title = ['Approach', 'PrecisionHT', 'RecallHT', 'F1HT', '#HT', 'PrecisionLT', 'RecallLT', 'F1LT', '#LT']
    DL = ['iTAPE+DL', precisionHT_DL, recallHT_DL, HT_F1_DL, num_DL_res, precisionLT_DL, recallLT_DL, LT_F1_DL,
          len(score) - num_DL_res]
    IR = ['iTAPE+IR', precisionHT_IR, recallHT_IR, HT_F1_IR, num_IR_res, precisionLT_IR, recallLT_IR, LT_F1_IR,
          len(score) - num_IR_res]
    both = ['iTAPE+tool', precisionHT_both, recallHT_both, HT_F1_both, num_both_res, precisionLT_both, recallLT_both,
            LT_F1_both, len(score) - num_both_res]
    sheet1.append(row_title)
    sheet1.append(DL)
    sheet1.append(IR)
    sheet1.append(both)


    result.save('./result_last/result_' + model_name + '.xlsx')